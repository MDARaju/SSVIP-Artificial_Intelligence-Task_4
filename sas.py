import face_recognition
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime

# Load images from the train folder
train_folder = 'train'
train_images = []
train_labels = []

for filename in os.listdir(train_folder):
    image_path = os.path.join(train_folder, filename)
    image = face_recognition.load_image_file(image_path)
    encoding = face_recognition.face_encodings(image)[0]
    train_images.append(encoding)
    # Extract person's name from the filename (assuming the filename is personX.jpg)
    person_name = os.path.splitext(filename)[0].replace("person", "")
    train_labels.append(person_name)

# Load the attendance Excel file or create a new one if it doesn't exist
attendance_file = 'attendance.xlsx'
wb = None
if os.path.exists(attendance_file):
    wb = load_workbook(attendance_file)
else:
    wb = Workbook()
    ws = wb.active
    # Write header row
    ws.append(['Date'] + train_labels)  # Column A for Date, rest for person names

# Get the current date
current_date = datetime.now().strftime("%Y-%m-%d")

# Check if the current date already exists in the Excel sheet
date_exists = False

# Initialize ws here
ws = wb.active

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] == current_date:
        date_exists = True
        break

# Create a dictionary to track whether each person is recognized
recognized_dict = {label: 0 for label in train_labels}

# Load test images with multiple faces from the test folder
test_folder = 'test'
for filename in os.listdir(test_folder):
    image_path = os.path.join(test_folder, filename)
    image = face_recognition.load_image_file(image_path)
    face_locations = face_recognition.face_locations(image)
    face_encodings = face_recognition.face_encodings(image, face_locations)

    # Iterate through detected faces in the test image
    for face_encoding in face_encodings:
        # Compare the face with the known faces
        results = face_recognition.compare_faces(train_images, face_encoding)

        # Check if True exists in the results list
        if True in results:
            for label, result in zip(train_labels, results):
                if result:
                    recognized_dict[label] = 1

# If the date doesn't exist, append a new row with recognition results  
if not date_exists:
    ws.append([current_date] + [recognized_dict[label] for label in train_labels])

# Save the attendance to the Excel file
wb.save(attendance_file)
print(f"Attendance has been recorded in {attendance_file}")
